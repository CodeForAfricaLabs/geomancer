from flask import current_app
from pickle import loads, dumps
from redis import Redis
from uuid import uuid4
import sys
import os
from cStringIO import StringIO
from csvkit.unicsv import UnicodeCSVReader, UnicodeCSVWriter
from geomancer.mancers.base import MancerError
from geomancer.helpers import import_class
from geomancer.app_config import RESULT_FOLDER, MANCERS
from datetime import datetime
import xlwt
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from itertools import izip_longest

redis = Redis()

try:
    from raven import Client
    from geomancer.app_config import SENTRY_DSN
    client = Client(dsn=SENTRY_DSN)
except ImportError:
    client = None
except KeyError:
    client = None

class DelayedResult(object):
    def __init__(self, key):
        self.key = key
        self._rv = None

    @property
    def return_value(self):
        if self._rv is None:
            rv = redis.get(self.key)
            if rv is not None:
                self._rv = loads(rv)
        return self._rv
    
def queuefunc(f):
    def delay(*args, **kwargs):
        qkey = current_app.config['REDIS_QUEUE_KEY']
        key = '%s:result:%s' % (qkey, str(uuid4()))
        s = dumps((f, key, args, kwargs))
        redis.rpush(current_app.config['REDIS_QUEUE_KEY'], s)
        return DelayedResult(key)
    f.delay = delay
    return f

@queuefunc
def do_the_work(file_contents, field_defs, filename):
    """
      field_defs looks like:
      {
        10: {
          'type': 'city_state', 
          'append_columns': ['total_population', 'median_age']
        }
      }

      file_contents is a string containing the contents of the uploaded file.
    """
    contents = StringIO(file_contents)
    reader = UnicodeCSVReader(contents)
    header = reader.next()
    result = None
    geo_ids = set()
    mancer_mapper = {}
    for mancer in MANCERS:
        m = import_class(mancer)()
        mancer_cols = [k['table_id'] for k in m.column_info()]
        for k, v in field_defs.items():
            field_cols = v['append_columns']
            for f in field_cols:
                if f in mancer_cols:
                    mancer_mapper[f] = {
                        'mancer': m,
                        'geo_id_map': {},
                        'geo_ids': set(),
                        'geo_type': v['type']
                    }
    for row_idx, row in enumerate(reader):
        col_idxs = [int(k) for k in field_defs.keys()]
        for idx in col_idxs:
            val = row[idx]
            geo_type = field_defs[idx]['type']
            for column in field_defs[idx]['append_columns']:
                mancer = mancer_mapper[column]['mancer']
                try:
                    if val:
                        geoid_search = mancer.geo_lookup(val, geo_type=geo_type)
                    else:
                        continue
                except MancerError, e:
                    return 'Error message: %s, Body: %s' % (e.message, e.body)
                row_geoid = geoid_search['geoid']
                if row_geoid:
                    mancer_mapper[column]['geo_ids'].add(row_geoid)
                    try:
                        mancer_mapper[column]['geo_id_map'][row_geoid].append(row_idx)
                    except KeyError:
                        mancer_mapper[column]['geo_id_map'][row_geoid] = [row_idx]
    all_data = {'header': []}
    output = []
    contents.seek(0)
    all_rows = list(reader)
    included_idxs = set()
    header_row = all_rows.pop(0)

    response = {
        'download_url': None,
        'geo_col': field_defs[0]['type'],
        'num_rows': len(all_rows),
        'num_matches': 0,
        'num_missing': 0,
        'cols_added': header_row[:]
    }

    for column, defs in mancer_mapper.items():
        geo_ids = defs['geo_ids']
        all_data.update({gid:[] for gid in geo_ids})
        geoid_mapper = defs['geo_id_map']
        geo_type = defs['geo_type']
        if geo_ids:
            mancer = defs['mancer']
            try:
                gids = [(geo_type, g,) for g in list(geo_ids)]
                data = mancer.search(geo_ids=gids, columns=[column])
            except MancerError, e:
                if client:
                    client.captureException()
                raise e
            all_data['header'].extend(data['header'])
            for gid in geo_ids:
                all_data[gid].extend(data[gid])
        else:
            raise MancerError('No geographies matched')
        for col in all_data['header']:
            if col not in header_row:
                header_row.append(col)
        i = 1
        try:
            output[0] = header_row
        except IndexError:
            output.append(header_row)
        for geoid, row_ids in geoid_mapper.items():
            for row_id in row_ids:
                included_idxs.add(row_id)
                row = all_rows[row_id]
                row.extend(all_data[geoid])
                try:
                    output[i] = row
                    i += 1
                except IndexError:
                    output.append(row)
        all_row_idxs = set(list(range(len(all_rows))))
        missing_rows = all_row_idxs.difference(included_idxs)
        response['num_missing'] = len(missing_rows) # store away missing rows
        for idx in missing_rows:
            row = all_rows[idx]
            row.extend(['' for i in header])
            output.append(row)
    name, ext = os.path.splitext(filename)
    fname = '%s_%s%s' % (name, datetime.now().isoformat(), ext)
    fpath = '%s/%s' % (RESULT_FOLDER, fname)
    if ext == '.xlsx':
        writeXLSX(fpath, output)
    elif ext == '.xls':
        writeXLS(fpath, output)
    else:
        writeCSV(fpath, output)

    response['download_url'] = '/download/%s' % fname
    response['num_matches'] = response['num_rows'] - response['num_missing']
    response['cols_added'] = list(set(header_row) - set(response['cols_added']))

    return response

def writeXLS(fpath, output):
    with open(fpath, 'wb') as f:
        workbook = xlwt.Workbook(encoding='utf-8')
        sheet = workbook.add_sheet('Geomancer Output')
        for r, row in enumerate(output):
            for c, col in enumerate(output[0]):
                sheet.write(r, c, row[c])
        workbook.save(fpath)

def writeXLSX(fpath, output):
    with open(fpath, 'wb') as f:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Geomancer Output'
        numcols = len(output[0])
        for r, row in enumerate(output):
            sheet.append([row[col_idx] for col_idx in range(numcols)])
        workbook.save(fpath)

def writeCSV(fpath, output):
    with open(fpath, 'wb') as f:
        writer = UnicodeCSVWriter(f)
        writer.writerows(output)

def queue_daemon(app, rv_ttl=500):
    print 'Mancing commencing...'
    while 1:
        msg = redis.blpop(app.config['REDIS_QUEUE_KEY'])
        func, key, args, kwargs = loads(msg[1])
        try:
            rv = func(*args, **kwargs)
            rv = {'status': 'ok', 'result': rv}
        except Exception, e:
            if client:
                client.captureException()
            rv = {'status': 'error', 'result': e.message}
        if rv is not None:
            redis.set(key, dumps(rv))
            redis.expire(key, rv_ttl)